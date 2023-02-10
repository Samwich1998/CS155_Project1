#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Need to Install on the Anaconda Prompt:
    $ pip install pyexcel
"""


# Basic Modules
import os
import sys
import numpy as np
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
# Openpyxl Styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font


class dataProcessing:        
        
    def xls2xlsx(self, excelFile, outputFolder):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(excelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return excelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        os.makedirs(outputFolder, exist_ok = True)
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(excelFile)
        newExcelFile = outputFolder + filename + "x"
        pyexcel.save_as(file_name = excelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Return New Excel File Name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to See if CSV Conversion Alreayd Occurred
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        """
        inputFile: The Input TXT/CSV File to Convert XLSX
        excelFile: The Output Excel File Name (XLSX)
        """
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            # Make Excel WorkBook
            xlWorkbook = xl.Workbook()
            xlWorksheet = xlWorkbook.active
            # Write the Data from the CSV File to the Excel WorkBook
            with open(inputFile, "r") as inputData:
                inReader = csv.reader(inputData, delimiter = excelDelimiter)
                with open(excelFile, 'w+', newline=''):
                    for row in inReader:
                        xlWorksheet.append(row)
            # Save as New Excel File
            xlWorkbook.save(excelFile)
        # Else Load the Data from the Excel File
        else:
            # Load the Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheet
    

class saveData():
    
    def addExcelAesthetics(self, WB_worksheet):
        # Center the Data in the Cells
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)        
        for column_cells in WB_worksheet.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            WB_worksheet.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length
            
            for cell in column_cells:
                cell.alignment = align
        # Header Style
        for cell in WB_worksheet["1:1"]:
            cell.font = Font(color='00FF0000', italic=True, bold=True)
        
        return WB_worksheet
    
    def saveData(self, dataToSave, saveDataFolder, saveExcelName, sheetName = "UV-Vis Analysis"):
        print("Saving the Data")
        # Create Output File Directory to Save Data: If Not Already Created
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Create Path to Save the Excel File
        excelFile = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create it
        if not os.path.isfile(excelFile):
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        else:
            print("Excel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile)
            WB_worksheet = WB.create_sheet(sheetName)
            
        # Add the Header
        headers = ["Sample Name", "Wavelength (nm)", "Absorbance (AU)", "Baseline Subtracted (AU)"]
        WB_worksheet.append(headers)
        
        # Organize and save the data
        for data in dataToSave:
            # Write the Data to Excel
            WB_worksheet.append(data)
        
        # Add Excel Aesthetics
        WB_worksheet = self.addExcelAesthetics(WB_worksheet)    
            
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()
    

class processFiles(dataProcessing):
    
    def extractData(self, filePath):
        print("Extracting data from file:", filePath)
        
        fileData = []
        # reading csv file
        with open(filePath, 'r') as csvfile:
            # creating a csv reader object
            csvreader = csv.reader(csvfile)
             
            # extracting field names through first row
            featureNames = np.array(next(csvreader)[2:])
         
            # extracting each data row one by one
            for row in csvreader:
                fileData.append(row)
         
            # get total number of rows
            print("\tTotal no. of rows: %d"%(csvreader.line_num))
        fileData = np.array(fileData)
        
        
        Y = np.array(fileData[:, 1])
        X = np.array(fileData[:, 2:], dtype=float)
        filenames = np.array(fileData[:, 0])
        
        return X, Y, featureNames, filenames
    
